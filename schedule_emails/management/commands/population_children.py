import os
from core import settings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from django.db import connection
from django.contrib.auth.models import User
from groups.models import Group
from schedule_emails.methods import enviar_correo
from openpyxl.styles import Alignment, Font, Border, Side
from django.core.management.base import BaseCommand, CommandError
from openpyxl.utils import get_column_letter

class Command(BaseCommand):

    def create_report(self, wb):
        heading = ('region','grupo','username','first_name','last_name','email','familias','ultima_semana_activo',)
        ws = wb.create_sheet('Reporte')
        ws.append(heading)

        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )

        cursor = connection.cursor()

        query = """
            select
                group_concat(distinct grupo_padre.name order by grupo_padre.name asc) as 'region',
                group_concat(distinct groups_group.name order by groups_group.name asc) as 'grupo',
                auth_user.username,
                auth_user.first_name,
                auth_user.last_name,
                auth_user.email,
                count(distinct groups_assignationmessengeruser.user_id ) as 'familias',
                if(week(auth_user.last_login) >= week(now())-1, 'Si', 'No') as 'ultima_semana_activo'
            from auth_user
                inner join groups_rolegroupuser
                    on groups_rolegroupuser.user_id = auth_user.id
                inner join groups_group
                    on groups_rolegroupuser.group_id = groups_group.id
                inner join groups_group as grupo_padre
                    on groups_group.parent_id = grupo_padre.id
                left join groups_assignationmessengeruser
                    on groups_assignationmessengeruser.group_id = groups_group.id
            where grupo_padre.parent_id = 38
            group by auth_user.id
            order by count(grupo_padre.id) asc, count(groups_group.id) asc, grupo_padre.name asc, groups_group.name asc, username asc
        """
        cursor.execute(query)
        result = cursor.fetchall()

        if len(result) > 0:
            for row in result:
                ws.append(row)

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            for i, col in enumerate(ws.columns):
                ws.column_dimensions[get_column_letter(i+1)].width = 25

    def query_detalle(self, tipo = 'ingreso'):
        query = """
            select
            count(distinct groups_assignationmessengeruser.user_id) as 'familias'
            from groups_group
                left join groups_group as grupo_padre
                    on groups_group.parent_id = grupo_padre.id
                left join groups_assignationmessengeruser
                    on groups_assignationmessengeruser.group_id = groups_group.id
                left join user_sessions_interaction
                    on user_sessions_interaction.user_id = groups_assignationmessengeruser.user_id
                    and user_sessions_interaction.`type` = 'session_init'
                left join posts_interaction
                    on posts_interaction.user_id = groups_assignationmessengeruser.user_id
                    and posts_interaction.`type` = 'dispatched'
                left join groups_rolegroupuser
                    on groups_rolegroupuser.group_id = groups_group.id
                left join auth_user
                    on groups_rolegroupuser.user_id = auth_user.id
            where grupo_padre.id in( select id from groups_group where parent_id = 38)
            and grupo_padre.id = %s
            group by groups_group.id
        """
        if tipo == 'ingreso':
            query += """ having familias > 0"""
        else:
            query += """ having familias = 0"""

        query += """ order by grupo_padre.name asc, groups_group.name asc"""

        return query

    def query_general(self):
        return """
        select grupo_padre.name as 'Región',
            groups_group.name as 'Grupo',
            grupo_padre.id,
            count(distinct groups_assignationmessengeruser.user_id) as 'familias',
            count(distinct posts_interaction.id) as 'actividades',
            count(distinct user_sessions_interaction.id) as 'sesiones',
            count(distinct case when auth_user.last_login is null
                    then NULL else auth_user.id end) as 'Ingresaron',
            count(distinct case when auth_user.last_login is null
                    then auth_user.id else NULL end) as 'Pendientes'
        from groups_group
            left join groups_group as grupo_padre
                on groups_group.parent_id = grupo_padre.id
            left join groups_assignationmessengeruser
                on groups_assignationmessengeruser.group_id = groups_group.id
            left join user_sessions_interaction
                on user_sessions_interaction.user_id = groups_assignationmessengeruser.user_id
                and user_sessions_interaction.`type` = 'session_init'
            left join posts_interaction
                on posts_interaction.user_id = groups_assignationmessengeruser.user_id
                and posts_interaction.`type` = 'dispatched'
            left join groups_rolegroupuser
                on groups_rolegroupuser.group_id = groups_group.id
            left join auth_user
                on groups_rolegroupuser.user_id = auth_user.id
        where grupo_padre.id in( select id from groups_group where parent_id = 38)
        group by grupo_padre.id
        order by grupo_padre.name asc, groups_group.name asc
        """

    def handle(self, *args, **kwargs):
        archivo = os.path.join(settings.BASE_DIR, 'schedule_emails/reporte.xlsx')
        wb = Workbook()
        std=wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(std)
        ws = wb.create_sheet('Consolidado')

        cursor = connection.cursor()
        cursor.execute(self.query_general())
        result = cursor.fetchall()

        ingreso = sum(row[6] for row in result)
        pendiente = sum(row[7] for row in result)
        total_general_1 = ingreso + pendiente
        porcentaje_general_1 = (ingreso * 100) / total_general_1

        if len(result) > 0:
            totales = [
                ('',),
                ('TOTALES GENERALES',),
                ('',),
                ('SOBRE USO DE FAMILIAS',),
                ('Total familias',sum(row[2] for row in result)),
                ('Actividades totales', sum(row[3] for row in result)),
                ('Total de sesiones', sum(row[4] for row in result)),
                ('Total niños con riesgos identificados', 0),
                ('',),
                ('SOBRE USO DE PROFESIONALES',),
                ('','Ingresaron','Pendientes','% Ingreso'),
                ('Profesionales que han ingresado',ingreso,pendiente,f"{porcentaje_general_1:.2f}"),
                ('Grupos que han ingresado familias',0,0,0),
                ('',),
                ('',)
            ]

            for fila in totales:
                ws.append(fila)

            acum_ingreso = 0
            acum_pendiente = 0

            for i,row in enumerate(result):

                ws.cell(row=ws.max_row+1, column=1, value="REGIÓN {0}".format(row[0]))

                query_ingresaron = self.query_detalle()
                cursor.execute(query_ingresaron,[row[2]])
                result1 = cursor.fetchall()

                query_pendientes = self.query_detalle('pendiente')
                cursor.execute(query_pendientes,[row[2]])
                result2 = cursor.fetchall()

                total_profesionales = row[6] + row[7]
                porcentaje_profesionales = (row[6] * 100) / total_profesionales
                total_grupo = len(result1) + len(result2)
                porcentaje_grupo = (len(result1) * 100) / total_grupo

                data_total = [
                    ('',),
                    ('SOBRE USO DE FAMILIAS',),
                    ('Total familias',row[2]),
                    ('Actividades totales', row[3]),
                    ('Total de sesiones', row[4]),
                    ('Total niños con riesgos identificados', 0),
                    ('',),
                    ('',),
                    ('SOBRE USO DE PROFESIONALES',),
                    ('','Ingresaron','Pendientes','% Ingreso'),
                    ('Profesionales que han ingresado',row[6],row[7],f"{porcentaje_profesionales:.2f}"),
                    ('Grupos que han ingresado familias',len(result1),len(result2),f"{porcentaje_grupo:.2f}"),
                    ('',),
                    ('',)
                ]

                acum_ingreso = acum_ingreso + len(result1)
                acum_pendiente = acum_pendiente + len(result2)

                for fila in data_total:
                    ws.append(fila)

            ws['b13'].value = acum_ingreso
            ws['c13'].value = acum_pendiente
            t = (acum_ingreso * 100) / (acum_ingreso+acum_pendiente)
            ws['d13'].value = f"{t:.2f}"

            for i, col in enumerate(ws.columns):
                ws.column_dimensions[get_column_letter(i+1)].width = 35

            for row in ws["B1":"D100"]:
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")

            self.create_report(wb)

            wb.save(archivo)

            query = """
            select distinct(email), first_name, last_name from auth_user
                inner join groups_rolegroupuser
                    on groups_rolegroupuser.user_id = auth_user.id
                inner join groups_group
                    on groups_rolegroupuser.group_id = groups_group.id
                inner join groups_group as grupo_padre
                    on groups_group.parent_id = grupo_padre.id
                left join groups_assignationmessengeruser
                    on groups_assignationmessengeruser.group_id = groups_group.id
            where grupo_padre.parent_id = 38
            and email is not null
            and email != ''
            """
            cursor.execute(query)
            result = cursor.fetchall()
            if len(result) > 0:
                for user in result:
                    if user[0]:
                        # enviar_correo(asunto='Población de niños',template='schedule_emails/population_children.html',data={
                        #     'user': {
                        #         'first_name': user[1],
                        #         'last_name': user[2]
                        #     }
                        # }, recipients=[user[0]], attachment_file=archivo)

                        enviar_correo(asunto='Población de niños',template='schedule_emails/population_children.html',data={
                            'user': {
                                'first_name': user[1],
                                'last_name': user[2]
                            }
                        }, recipients=['alejandro.reyna@afinidata.com','lgodoy@afinidata.com','ac@afinidata.com'], attachment_file=archivo)

